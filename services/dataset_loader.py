from pathlib import Path

from openpyxl import load_workbook

from config import INTERESTS, MOODS
from services.genre_service import get_genres

GENRE_KEYWORDS = {
  "Фантастика": ["фантастик", "sci-fi"],
  "Детектив": ["детектив", "триллер", "криминал", "шпион"],
  "Роман": ["роман", "проза", "драма", "повесть"],
  "Ужасы": ["ужас", "хоррор", "мистика"],
  "Фэнтези": ["фэнтези", "сказк"],
  "Биография": ["биограф", "мемуар", "автобиограф"],
  "Научпоп": ["научпоп", "популярн"],
  "Классика": ["классик", "эпопея", "поэма"],
  "Приключения": ["приключен", "путешеств"],
  "Поэзия": ["поэз", "стих", "поэма"],
}

INTEREST_KEYWORDS = {
  "История": ["истори", "историч"],
  "Философия": ["философ", "утопи"],
  "Психология": ["психолог"],
  "Наука": ["науч", "наукоп"],
  "Любовь": ["любов", "романт"],
  "Приключения": ["приключен"],
  "Мистика": ["мистик", "оккульт"],
  "Юмор": ["юмор", "сатир", "комед"],
}


def load_books_from_xlsx(file_path):
  path = Path(file_path)
  if not path.exists():
    raise FileNotFoundError(f"Файл датасета не найден: {path}")

  workbook = load_workbook(path, read_only=True, data_only=True)
  worksheet = workbook.active
  rows = list(worksheet.iter_rows(values_only=True))
  workbook.close()

  if not rows:
    return []

  headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
  column_map = _build_column_map(headers)
  books = []

  for row in rows[1:]:
    if not row or not row[column_map["title"]]:
      continue

    title = str(row[column_map["title"]]).strip()
    author = str(row[column_map["author"]]).strip() if row[column_map["author"]] else "Неизвестный автор"
    genre_raw = str(row[column_map["genre"]]).strip() if row[column_map["genre"]] else ""
    pages = _parse_pages(row[column_map["pages"]])
    rating = _parse_rating(row[column_map["rating"]])

    books.append({
      "title": title,
      "author": author,
      "genre": genre_raw or _normalize_genre(genre_raw),
      "mood": _pick_mood(title),
      "description": _build_description(title, author, genre_raw),
      "rating": rating,
      "pages": pages,
      "is_new": 0,
      "interests": _extract_interests(genre_raw),
    })

  return books


def _build_column_map(headers):
  title_idx = _find_column(headers, ("title", "название"))
  author_idx = _find_column(headers, ("author", "автор"))
  genre_idx = _find_column(headers, ("жанр", "genre"))
  pages_idx = _find_column(headers, ("кол_страниц", "страниц", "pages"))
  rating_idx = _find_column(headers, ("rating", "рейтинг"))

  missing = []
  if title_idx is None:
    missing.append("title")
  if author_idx is None:
    missing.append("author")
  if genre_idx is None:
    missing.append("жанр")
  if pages_idx is None:
    missing.append("кол_страниц")
  if rating_idx is None:
    missing.append("rating")

  if missing:
    raise ValueError(f"В xlsx не найдены колонки: {', '.join(missing)}")

  return {
    "title": title_idx,
    "author": author_idx,
    "genre": genre_idx,
    "pages": pages_idx,
    "rating": rating_idx,
  }


def _find_column(headers, names):
  for index, header in enumerate(headers):
    if header in names:
      return index
  return None


def _parse_pages(value):
  if value is None or value == "":
    return 300
  try:
    return int(float(str(value).replace(",", ".").strip()))
  except ValueError:
    return 300


def _parse_rating(value):
  try:
    rating = float(str(value).replace(",", ".").strip())
  except (TypeError, ValueError):
    rating = 0
  return rating if rating > 0 else 4.0


def _normalize_genre(genre_raw):
  lower = genre_raw.lower()
  for genre in get_genres():
    if genre.lower() in lower:
      return genre
  for genre, keywords in GENRE_KEYWORDS.items():
    if any(keyword in lower for keyword in keywords):
      return genre
  return genre_raw.strip() or "Классика"


def _extract_interests(genre_raw):
  lower = genre_raw.lower()
  found = []
  for interest in INTERESTS:
    if interest.lower() in lower:
      found.append(interest)
  for interest, keywords in INTEREST_KEYWORDS.items():
    if interest in found:
      continue
    if any(keyword in lower for keyword in keywords):
      found.append(interest)
  return ",".join(found)


def _pick_mood(title):
  return MOODS[hash(title) % len(MOODS)]


def _build_description(title, author, genre_raw):
  if genre_raw:
    return f"«{title}» — {author}. Жанры: {genre_raw}."
  return f"«{title}» — {author}."
