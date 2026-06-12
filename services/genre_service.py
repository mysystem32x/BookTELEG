from pathlib import Path

from openpyxl import load_workbook

from config import DATASET_PATH

GENRES_PER_PAGE = 10

_genres_cache = None


def get_genres():
  global _genres_cache
  if _genres_cache is None:
    _genres_cache = load_genres_from_dataset(DATASET_PATH)
  return _genres_cache


def load_genres_from_dataset(file_path):
  path = Path(file_path)
  if not path.exists():
    return []

  workbook = load_workbook(path, read_only=True, data_only=True)
  worksheet = workbook.active
  rows = list(worksheet.iter_rows(values_only=True))
  workbook.close()

  if not rows:
    return []

  headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
  genre_idx = _find_column(headers, ("жанр", "genre"))
  if genre_idx is None:
    return []

  genres = set()
  for row in rows[1:]:
    if not row or not row[genre_idx]:
      continue
    for part in str(row[genre_idx]).split(","):
      genre = part.strip()
      if genre:
        genres.add(genre)

  return sorted(genres)


def split_book_genres(genre_str):
  if not genre_str:
    return []
  return [part.strip() for part in str(genre_str).split(",") if part.strip()]


def genre_matches(book_genre, user_genres):
  book_tokens = {genre.lower() for genre in split_book_genres(book_genre)}
  return any(genre.strip().lower() in book_tokens for genre in user_genres)


def matched_genres(book_genre, user_genres):
  book_tokens = {genre.lower() for genre in split_book_genres(book_genre)}
  return [genre for genre in user_genres if genre.strip().lower() in book_tokens]


def _find_column(headers, names):
  for index, header in enumerate(headers):
    if header in names:
      return index
  return None
